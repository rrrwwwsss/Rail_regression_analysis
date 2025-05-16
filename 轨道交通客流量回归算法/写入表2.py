import openpyxl
def seve_file(result3,row):
    # 加载现有的 Excel 文件
    file_path = './结果/师兄SHAP指标分析_new.xlsx'  # 替换为你的文件路径
    wb = openpyxl.load_workbook(file_path)

    # 选择活动工作表
    ws = wb.active

    # 示例字典
    feature_dict = result3

    # 获取第一列（A列）的特征名称
    feature_names = [ws.cell(row=i, column=1).value for i in range(3, 40)]

    # 遍历字典，将值插入到对应行的第二列
    for i, feature in enumerate(feature_names, start=3):
        if feature in feature_dict:
            ws.cell(row=i, column=row, value=feature_dict[feature])

    # 保存修改后的 Excel 文件
    wb.save(file_path)  # 替换为你想保存的文件路径
